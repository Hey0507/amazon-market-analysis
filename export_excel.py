"""
export_excel.py - 市场分析报告 Excel 导出脚本
运行: python export_excel.py
"""
import pandas as pd
import pickle
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

# ── 颜色常量 ───────────────────────────────────────────────
BLUE_DARK   = "1E3A5F"
BLUE_MID    = "2563EB"
BLUE_LIGHT  = "DBEAFE"
GRAY_HEADER = "F1F5F9"
GRAY_ROW    = "F8FAFC"
WHITE       = "FFFFFF"
RED_ACCENT  = "E11D48"
GREEN_ACC   = "10B981"
BORDER_COL  = "CBD5E1"

def make_border():
    s = Side(style='thin', color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(ws, row, col, value, bg=BLUE_DARK, fg=WHITE, bold=True, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()
    return cell

def data_style(ws, row, col, value, bg=WHITE, bold=False, num_fmt=None, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, color="1E293B", size=10, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = make_border()
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, text, row, ncols, bg=BLUE_DARK):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

# ══════════════════════════════════════════════════════════════
# 加载数据
# ══════════════════════════════════════════════════════════════
print("📂 加载数据...")
df = pd.read_pickle("processed_market_data.pkl")
df['销量'] = pd.to_numeric(df['销量'], errors='coerce')
df['销售额'] = pd.to_numeric(df['销售额'], errors='coerce')
df['月份_dt'] = pd.to_datetime(df['月份'])

# 最新完整季度
df['季度'] = df['月份_dt'].dt.to_period('Q')
q_counts = df.groupby('季度')['月份'].nunique()
complete_qs = q_counts[q_counts == 3].index
latest_q = complete_qs.max() if not complete_qs.empty else df['季度'].max()
q_df = df[df['季度'] == latest_q]

price_order = ['0-30', '30-50', '50-100', '100-150', '150+']

wb = Workbook()

# ══════════════════════════════════════════════════════════════
# Sheet 1: 市场概览
# ══════════════════════════════════════════════════════════════
print("📊 Sheet 1: 市场概览...")
ws1 = wb.active
ws1.title = "市场概览"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 40
ws1.row_dimensions[2].height = 20

title_row(ws1, f"Amazon 无线麦克风市场分析报告  |  数据区间: {df['月份'].min()} ~ {df['月份'].max()}  |  生成时间: {datetime.now().strftime('%Y-%m-%d')}", 1, 6)

# KPI 卡片
kpis = [
    ("总销量", f"{int(df['销量'].sum()):,}", BLUE_MID),
    ("总销售额 (USD)", f"${df['销售额'].sum()/1e6:.1f}M", BLUE_MID),
    ("覆盖 ASIN 数", str(df['ASIN'].nunique()), BLUE_MID),
    ("市场均价 (USD)", f"${df['价格'].mean():.2f}", BLUE_MID),
    ("数据月份数", str(df['月份'].nunique()), BLUE_MID),
    ("覆盖品牌数", str(df['品牌'].nunique()), BLUE_MID),
]

ws1.merge_cells("A3:F3")
ws1.cell(row=3, column=1, value="📌 核心 KPI 指标").font = Font(bold=True, size=11, color=BLUE_DARK)

for i, (label, val, color) in enumerate(kpis, 1):
    col = i
    ws1.row_dimensions[4].height = 18
    ws1.row_dimensions[5].height = 32
    ws1.row_dimensions[6].height = 18
    header_style(ws1, 4, col, label, bg=GRAY_HEADER, fg=BLUE_DARK, bold=True, size=10)
    cell = ws1.cell(row=5, column=col, value=val)
    cell.fill = PatternFill("solid", fgColor=BLUE_LIGHT)
    cell.font = Font(bold=True, size=14, color=BLUE_DARK, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = make_border()

set_col_widths(ws1, [18, 18, 18, 18, 18, 18])

# ── 月度趋势表 ──
ws1.merge_cells("A8:F8")
ws1.cell(row=8, column=1, value="📈 月度销售趋势 (全市场 vs Hollyland)").font = Font(bold=True, size=11, color=BLUE_DARK)
ws1.row_dimensions[8].height = 22

hdrs = ["月份", "全市场销量", "全市场销售额 (USD)", "Hollyland销量", "Hollyland销售额 (USD)", "Hollyland占比"]
for j, h in enumerate(hdrs, 1):
    header_style(ws1, 9, j, h, bg=BLUE_DARK)

total_trend = df.groupby('月份')[['销量','销售额']].sum().reset_index().sort_values('月份')
hl_trend = df[df['品牌']=='Hollyland'].groupby('月份')[['销量','销售额']].sum().reset_index()
trend = total_trend.merge(hl_trend, on='月份', how='left', suffixes=('_全市场','_Hollyland')).fillna(0)
trend['占比'] = trend['销售额_Hollyland'] / trend['销售额_全市场']

for r_idx, row in trend.iterrows():
    r = 10 + list(trend.index).index(r_idx)
    bg = GRAY_ROW if (r % 2 == 0) else WHITE
    data_style(ws1, r, 1, row['月份'], bg=bg)
    data_style(ws1, r, 2, int(row['销量_全市场']), bg=bg, num_fmt='#,##0')
    data_style(ws1, r, 3, row['销售额_全市场'], bg=bg, num_fmt='$#,##0')
    data_style(ws1, r, 4, int(row['销量_Hollyland']), bg=bg, num_fmt='#,##0')
    data_style(ws1, r, 5, row['销售额_Hollyland'], bg=bg, num_fmt='$#,##0')
    data_style(ws1, r, 6, row['占比'], bg=bg, num_fmt='0.0%')

# 折线图
last_r = 9 + len(trend)
chart = LineChart()
chart.title = "全市场 vs Hollyland 月度销售额趋势"
chart.style = 10
chart.y_axis.title = "销售额 (USD)"
chart.x_axis.title = "月份"
chart.height = 12
chart.width = 24

data_ref1 = Reference(ws1, min_col=3, min_row=9, max_row=last_r)
data_ref2 = Reference(ws1, min_col=5, min_row=9, max_row=last_r)
cats = Reference(ws1, min_col=1, min_row=10, max_row=last_r)

chart.add_data(data_ref1, titles_from_data=True)
chart.add_data(data_ref2, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.line.solidFill = BLUE_MID
chart.series[0].graphicalProperties.line.width = 20000
chart.series[1].graphicalProperties.line.solidFill = RED_ACCENT
chart.series[1].graphicalProperties.line.width = 20000

ws1.add_chart(chart, f"A{last_r + 2}")


# ══════════════════════════════════════════════════════════════
# Sheet 2: 品牌份额
# ══════════════════════════════════════════════════════════════
print("📊 Sheet 2: 品牌份额...")
ws2 = wb.create_sheet("品牌市场份额")
ws2.sheet_view.showGridLines = False
title_row(ws2, "Top 15 品牌市场份额分析（按销售额）", 1, 5)
set_col_widths(ws2, [5, 24, 16, 16, 14])

hdrs2 = ["排名", "品牌", "总销量", "总销售额 (USD)", "销售额占比"]
for j, h in enumerate(hdrs2, 1):
    header_style(ws2, 2, j, h)

brand_df = df.groupby('品牌')[['销量','销售额']].sum().sort_values('销售额', ascending=False).head(15).reset_index()
total_rev = df['销售额'].sum()
brand_df['占比'] = brand_df['销售额'] / total_rev

bar_colors = [BLUE_MID, RED_ACCENT, GREEN_ACC, "F59E0B", "8B5CF6",
              "06B6D4", "EC4899", "14B8A6", "F97316", "6366F1",
              "84CC16", "EF4444", "0EA5E9", "D946EF", "78716C"]

for i, row in brand_df.iterrows():
    r = 3 + i
    bg = BLUE_LIGHT if row['品牌'] == 'Hollyland' else (GRAY_ROW if i % 2 == 0 else WHITE)
    bold = row['品牌'] == 'Hollyland'
    data_style(ws2, r, 1, i+1, bg=bg, bold=bold)
    data_style(ws2, r, 2, row['品牌'], bg=bg, bold=bold, align="left")
    data_style(ws2, r, 3, int(row['销量']), bg=bg, num_fmt='#,##0')
    data_style(ws2, r, 4, row['销售额'], bg=bg, num_fmt='$#,##0')
    data_style(ws2, r, 5, row['占比'], bg=bg, num_fmt='0.0%', bold=bold)

# 柱状图
chart2 = BarChart()
chart2.type = "bar"
chart2.title = "Top 15 品牌销售额"
chart2.style = 10
chart2.y_axis.title = "品牌"
chart2.x_axis.title = "销售额 (USD)"
chart2.height = 16
chart2.width = 22

data_r = Reference(ws2, min_col=4, min_row=2, max_row=2+len(brand_df))
cats_r = Reference(ws2, min_col=2, min_row=3, max_row=2+len(brand_df))
chart2.add_data(data_r, titles_from_data=True)
chart2.set_categories(cats_r)
chart2.series[0].graphicalProperties.solidFill = BLUE_MID
ws2.add_chart(chart2, "G2")


# ══════════════════════════════════════════════════════════════
# Sheet 3: 价位段分析
# ══════════════════════════════════════════════════════════════
print("📊 Sheet 3: 价位段分析...")
ws3 = wb.create_sheet("价位段分析")
ws3.sheet_view.showGridLines = False
title_row(ws3, "各价位段市场结构分析", 1, 7)
set_col_widths(ws3, [14, 14, 16, 14, 12, 14, 14])

hdrs3 = ["价位段 (USD)", "销量", "销售额 (USD)", "销售额占比", "ASIN数", "均价", "Top5品牌集中度"]
for j, h in enumerate(hdrs3, 1):
    header_style(ws3, 2, j, h)

price_df = df.groupby('价格区间').agg(
    销量=('销量','sum'), 销售额=('销售额','sum'),
    ASIN数=('ASIN','nunique'), 均价=('价格','mean')
).reindex(price_order).dropna().reset_index()
price_df['占比'] = price_df['销售额'] / price_df['销售额'].sum()

seg_colors = [BLUE_MID, GREEN_ACC, "F59E0B", RED_ACCENT, "8B5CF6"]

for i, row in price_df.iterrows():
    r = 3 + i
    bg = GRAY_ROW if i % 2 == 0 else WHITE
    tier_total = df[df['价格区间']==row['价格区间']]['销售额'].sum()
    top5 = df[df['价格区间']==row['价格区间']].groupby('品牌')['销售额'].sum().sort_values(ascending=False).head(5).sum()
    conc = top5/tier_total if tier_total > 0 else 0
    data_style(ws3, r, 1, row['价格区间'], bg=bg, bold=True)
    data_style(ws3, r, 2, int(row['销量']), bg=bg, num_fmt='#,##0')
    data_style(ws3, r, 3, row['销售额'], bg=bg, num_fmt='$#,##0')
    data_style(ws3, r, 4, row['占比'], bg=bg, num_fmt='0.0%')
    data_style(ws3, r, 5, int(row['ASIN数']), bg=bg, num_fmt='#,##0')
    data_style(ws3, r, 6, row['均价'], bg=bg, num_fmt='$#,##0.00')
    data_style(ws3, r, 7, conc, bg=bg, num_fmt='0.0%')

# ── 最新完整季度品牌明细 ──
ws3.merge_cells("A10:G10")
ws3.cell(row=10, column=1, value=f"📊 各价位段 Top 5 品牌占比 — 最新完整季度: {latest_q}").font = Font(bold=True, size=11, color=BLUE_DARK)
ws3.row_dimensions[10].height = 22

hdrs4 = ["价位段", "排名", "品牌", "季度销量", "季度销售额 (USD)", "在该价位段占比", ""]
for j, h in enumerate(hdrs4, 1):
    header_style(ws3, 11, j, h)

r_start = 12
for tier in price_order:
    tier_df = q_df[q_df['价格区间']==tier]
    t_total = tier_df['销售额'].sum()
    brands = tier_df.groupby('品牌')[['销量','销售额']].sum().sort_values('销售额', ascending=False).head(5).reset_index()
    for k, brow in brands.iterrows():
        bg = BLUE_LIGHT if brow['品牌']=='Hollyland' else (GRAY_ROW if k%2==0 else WHITE)
        data_style(ws3, r_start, 1, tier, bg=bg, bold=True)
        data_style(ws3, r_start, 2, k+1, bg=bg)
        data_style(ws3, r_start, 3, brow['品牌'], bg=bg, align="left")
        data_style(ws3, r_start, 4, int(brow['销量']), bg=bg, num_fmt='#,##0')
        data_style(ws3, r_start, 5, brow['销售额'], bg=bg, num_fmt='$#,##0')
        data_style(ws3, r_start, 6, brow['销售额']/t_total if t_total>0 else 0, bg=bg, num_fmt='0.0%')
        r_start += 1


# ══════════════════════════════════════════════════════════════
# Sheet 4: Hollyland 专项
# ══════════════════════════════════════════════════════════════
print("📊 Sheet 4: Hollyland 专项...")
ws4 = wb.create_sheet("Hollyland 专项")
ws4.sheet_view.showGridLines = False
title_row(ws4, "Hollyland 品牌专项分析", 1, 6, bg=RED_ACCENT)
set_col_widths(ws4, [14, 14, 18, 18, 16, 16])

hl_monthly = df[df['品牌']=='Hollyland'].groupby('月份').agg(
    销量=('销量','sum'), 销售额=('销售额','sum'), ASIN数=('ASIN','nunique'), 均价=('价格','mean')
).reset_index().sort_values('月份')
total_by_month = df.groupby('月份')['销售额'].sum()
hl_monthly['市场占比'] = hl_monthly.apply(lambda r: r['销售额']/total_by_month[r['月份']] if r['月份'] in total_by_month else 0, axis=1)

hdrs5 = ["月份", "销量", "销售额 (USD)", "市场占比", "ASIN数", "月均价 (USD)"]
for j, h in enumerate(hdrs5, 1):
    header_style(ws4, 2, j, h, bg=RED_ACCENT)

for i, row in hl_monthly.iterrows():
    r = 3 + list(hl_monthly.index).index(i)
    bg = GRAY_ROW if r%2==0 else WHITE
    data_style(ws4, r, 1, row['月份'], bg=bg)
    data_style(ws4, r, 2, int(row['销量']), bg=bg, num_fmt='#,##0')
    data_style(ws4, r, 3, row['销售额'], bg=bg, num_fmt='$#,##0')
    data_style(ws4, r, 4, row['市场占比'], bg=bg, num_fmt='0.0%')
    data_style(ws4, r, 5, int(row['ASIN数']), bg=bg)
    data_style(ws4, r, 6, row['均价'], bg=bg, num_fmt='$#,##0.00')


# ══════════════════════════════════════════════════════════════
# Sheet 5: 原始明细
# ══════════════════════════════════════════════════════════════
print("📊 Sheet 5: 原始明细...")
ws5 = wb.create_sheet("原始明细 (Top200)")
ws5.sheet_view.showGridLines = False
top200 = df.sort_values('销售额', ascending=False).head(200)
cols_export = ['月份','品牌','ASIN','品类','价格区间','价格','销量','销售额','地区','BSR']
cols_export = [c for c in cols_export if c in top200.columns]

title_row(ws5, "原始数据明细（按销售额排名前 200 条）", 1, len(cols_export))
for j, h in enumerate(cols_export, 1):
    header_style(ws5, 2, j, h)

for i, row in top200[cols_export].reset_index(drop=True).iterrows():
    r = 3 + i
    bg = GRAY_ROW if i%2==0 else WHITE
    for j, col in enumerate(cols_export, 1):
        v = row[col]
        fmt = None
        if col == '销售额': fmt = '$#,##0'
        elif col == '价格': fmt = '$#,##0.00'
        elif col == '销量': fmt = '#,##0'
        data_style(ws5, r, j, v, bg=bg, num_fmt=fmt, align="left" if col in ['品牌','ASIN','品类'] else "center")

for j in range(1, len(cols_export)+1):
    ws5.column_dimensions[get_column_letter(j)].width = 16


# ══════════════════════════════════════════════════════════════
# 保存
# ══════════════════════════════════════════════════════════════
out_path = f"Amazon市场分析报告_{datetime.now().strftime('%Y%m%d')}.xlsx"
wb.save(out_path)
print(f"\n✅ 报告已生成: {out_path}")
print("   包含以下工作表:")
print("   - 市场概览（KPI + 月度趋势图）")
print("   - 品牌市场份额（Top 15 + 柱状图）")
print("   - 价位段分析（结构 + 季度品牌明细）")
print("   - Hollyland 专项（月度明细）")
print("   - 原始明细（Top 200）")
